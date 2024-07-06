from os import listdir, getcwd
from os.path import isfile, join, basename
import openpyxl

data_path = getcwd() + '\\xlsxdata'

def find_all_xl_files():
    global data_path

    files = [join(data_path, f) for f in listdir(data_path)] # Get each item in the current working directory
    files = [f for f in files if isfile(f)] # ignore folders
    files = [f for f in files if f.endswith('.xlsx')] # only get valid excel document filetypes
    files = [f for f in files if 'Template' not in f] #ignore template files
    files = [f for f in files if '~$' not in f] # ignore any versions that are "open files" (ususally hidden in explorer)

    return files

all_xl_files = find_all_xl_files()

print([f for f in listdir(data_path)])
print([basename(f) for f in all_xl_files])

dice_colums = ['B', 'C', 'D', 'E', 'F', 'G', 'H']
die_size_row = '2'
average_row = '3'
expected_average_row = '4'
max_count_row = '5'
min_count_row = '6'
expected_count_row = '7'
roll_range_start_row = '11'
column_to_index = {
                    'B': 0, #d20
                    'C': 1, #d12
                    'D': 2, #d%
                    'E': 3, #d10
                    'F': 4, #d8
                    'G': 5, #d6
                    'H': 6  #d4
                 }

def print_xl_files():
    i = 0
    for f in all_xl_files:
        i += 1
        print(i, ':', f)


def get_user_input():
    file_count = len(all_xl_files)
    while True:
        user_input = input("Which file do you want to open? ")

        try:
            user_selection = int(user_input)
        except ValueError:
            print('Error: not a valid input')
            continue

        if not user_selection <= file_count or not user_selection > 0:
            print('Error: not a valid input')
            continue
        
        break
    return all_xl_files[user_selection - 1]

def get_frequencies(ws, die_frequency):
    for column in dice_colums:

        for row in range(int(roll_range_start_row),ws.max_row+1):
            if ws.cell(row, column_to_index[column]+2).value is None:
                break;
        
        cells = ws[column + roll_range_start_row : column + str(row)]
        for cell in cells: 
            die_frequency[column_to_index[column]][cell[0].value - 1] += 1;

    # return die_frequency

def load_worksheet(file = -1):
    if file == -1:
        print_xl_files()
        selection = get_user_input()
    else:
        selection = all_xl_files[file]

    wb = openpyxl.load_workbook(selection, data_only=True) # data_only tag pulls values and not formulas
    ws = wb.active

    return ws

def get_frequencies_from_file(file:int = -1):
    ws = load_worksheet(file)

    die_frequency = [ # I hate the way ive done this, find a way to do it dynamically and effectively
                        [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0],
                        [0,0,0,0,0,0,0,0,0,0,0,0],
                        [0,0,0,0,0,0,0,0,0,0],
                        [0,0,0,0,0,0,0,0,0,0],
                        [0,0,0,0,0,0,0,0],
                        [0,0,0,0,0,0],
                        [0,0,0,0]
                    ]
    
    get_frequencies(ws, die_frequency)

    return die_frequency